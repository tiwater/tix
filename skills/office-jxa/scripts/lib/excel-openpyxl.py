#!/usr/bin/env python3
"""
excel-openpyxl.py — Cross-platform Excel backend using openpyxl.

Works on Linux, Windows, and macOS without Microsoft Excel installed.
Requires: pip install openpyxl

Usage: python3 excel-openpyxl.py <function> [args...]
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def require_openpyxl():
    if not HAS_OPENPYXL:
        print(json.dumps({
            "error": "openpyxl not installed. Run: pip install openpyxl"
        }))
        sys.exit(1)


def open_workbook(file_path):
    """Open an .xlsx file."""
    require_openpyxl()
    if not file_path:
        print(json.dumps({
            "error": "File path required on this platform (no active workbook support)"
        }))
        sys.exit(1)
    if not os.path.exists(file_path):
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)
    return openpyxl.load_workbook(file_path)


def readStructure(file_path=None):
    """Read workbook overview."""
    wb = open_workbook(file_path)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "rowCount": ws.max_row or 0,
            "columnCount": ws.max_column or 0,
        })
    print(json.dumps({
        "name": Path(file_path).name,
        "path": os.path.abspath(file_path),
        "sheetCount": len(sheets),
        "sheets": sheets,
    }, ensure_ascii=False))


def readRange(file_path=None, sheet_name=None, range_str=None, read_all=False):
    """Read a range of cells."""
    wb = open_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if read_all or range_str == "all":
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(v) if v is not None else "" for v in row])
    else:
        data = []
        for row in ws[range_str]:
            data.append([str(cell.value) if cell.value is not None else "" for cell in row])

    print(json.dumps({
        "sheet": ws.title,
        "range": range_str or "all",
        "rows": len(data),
        "cols": len(data[0]) if data else 0,
        "data": data,
    }, ensure_ascii=False))


def searchText(file_path=None, query=""):
    """Search for text across all sheets."""
    wb = open_workbook(file_path)
    matches = []
    query_lower = query.lower()

    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and query_lower in str(cell.value).lower():
                    matches.append({
                        "sheet": name,
                        "address": cell.coordinate,
                        "value": str(cell.value),
                    })
                    if len(matches) >= 50:
                        break
            if len(matches) >= 50:
                break
        if len(matches) >= 50:
            break

    print(json.dumps({
        "query": query,
        "matchCount": len(matches),
        "matches": matches,
    }, ensure_ascii=False))


def createWorkbook(save_path=None, sheet_names=None):
    """Create a new workbook."""
    require_openpyxl()
    if not save_path:
        print(json.dumps({"error": "save_path required (--save-as PATH)"}))
        sys.exit(1)

    wb = openpyxl.Workbook()
    if sheet_names:
        names = sheet_names.split(",") if isinstance(sheet_names, str) else sheet_names
        wb.active.title = names[0].strip()
        for name in names[1:]:
            wb.create_sheet(title=name.strip())

    wb.save(save_path)
    print(json.dumps({
        "created": True,
        "path": os.path.abspath(save_path),
        "sheets": len(wb.sheetnames),
    }))


def writeRange(file_path=None, sheet_name=None, start_cell="A1", data_json=""):
    """Write data to cells from a 2D JSON array."""
    wb = open_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = json.loads(data_json)

    # Parse start cell (e.g., "A1" → col=1, row=1)
    from openpyxl.utils.cell import coordinate_from_string
    col_letter, start_row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(col_letter)

    for r, row_data in enumerate(data):
        row_vals = row_data if isinstance(row_data, list) else [row_data]
        for c, val in enumerate(row_vals):
            ws.cell(row=start_row + r, column=start_col + c, value=val)

    wb.save(file_path)
    print(json.dumps({
        "written": True,
        "sheet": ws.title,
        "startCell": start_cell,
        "rows": len(data),
        "cols": len(data[0]) if data else 0,
    }))


def addSheet(file_path=None, sheet_name="Sheet"):
    """Add a new worksheet."""
    wb = open_workbook(file_path)
    wb.create_sheet(title=sheet_name)
    wb.save(file_path)
    print(json.dumps({
        "added": True,
        "name": sheet_name,
        "totalSheets": len(wb.sheetnames),
    }))


def setFormula(file_path=None, sheet_name=None, cell_address="A1", formula=""):
    """Set a cell formula."""
    wb = open_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    ws[cell_address] = formula
    wb.save(file_path)
    print(json.dumps({
        "set": True,
        "cell": cell_address,
        "formula": formula,
    }))


def formatRange(file_path=None, sheet_name=None, range_str="A1", format_json=""):
    """Format a range of cells."""
    wb = open_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    fmt = json.loads(format_json)

    font_kwargs = {}
    if "bold" in fmt:
        font_kwargs["bold"] = fmt["bold"]
    if "italic" in fmt:
        font_kwargs["italic"] = fmt["italic"]
    if "fontSize" in fmt:
        font_kwargs["size"] = fmt["fontSize"]
    if "fontColor" in fmt:
        font_kwargs["color"] = fmt["fontColor"]

    fill = None
    if "bgColor" in fmt:
        fill = PatternFill(start_color=fmt["bgColor"], end_color=fmt["bgColor"], fill_type="solid")

    font = Font(**font_kwargs) if font_kwargs else None

    for row in ws[range_str]:
        for cell in row:
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if "numberFormat" in fmt:
                cell.number_format = fmt["numberFormat"]

    wb.save(file_path)
    print(json.dumps({"formatted": True, "range": range_str}))


def saveWorkbook(file_path=None, save_as=None):
    """Save or copy workbook."""
    wb = open_workbook(file_path)
    target = save_as or file_path
    wb.save(target)
    print(json.dumps({"saved": True, "path": os.path.abspath(target)}))


# ── Dispatcher ──

FUNCTIONS = {
    "readStructure": readStructure,
    "readRange": readRange,
    "searchText": searchText,
    "createWorkbook": createWorkbook,
    "writeRange": writeRange,
    "addSheet": addSheet,
    "setFormula": setFormula,
    "formatRange": formatRange,
    "saveWorkbook": saveWorkbook,
}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": f"Usage: {sys.argv[0]} <function> [args...]"}))
        sys.exit(1)

    func_name = sys.argv[1]
    if func_name not in FUNCTIONS:
        print(json.dumps({"error": f"Unknown: {func_name}", "available": list(FUNCTIONS.keys())}))
        sys.exit(1)

    args = sys.argv[2:]
    args = [None if a == "null" else a for a in args]
    FUNCTIONS[func_name](*args)
